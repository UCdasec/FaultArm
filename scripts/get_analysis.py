import os, re
import argparse
from openpyxl import Workbook, load_workbook, worksheet


class GetAnalysis():
    def __init__(self, folder_path: str, workbook_path: str):
        self.folder_path = folder_path
        self.workbook_path = workbook_path + 'analysis.xlsx'
        self.workbook = None
        self.active_sheet = None
        self.labelling = load_workbook(os.path.join(os.path.dirname(self.workbook_path), 'Dataset_Labelling.xlsx'),
                                       read_only=True)
        self.active_labelling_sheet = self.labelling['C vs Assembly cross-check']

    def createWorkbook(self, filepath: str) -> bool:
        try:
            # Check if the file already exists
            if os.path.exists(filepath):
                print(f"The workbook '{filepath}' already exists.")
                return True

            self.workbook = Workbook()
            self.workbook.save(filepath)

            print(f"The workbook is '{filepath}'")
            return True

        except Exception as e:
            print(f"Error creating the workbook: {e}")
            return False

    def generate_resultsTable(self, top_row: int):
        try:
            # Insert the results table
            pattern_headers = ["Pattern", "Branch", "Constant Coding", "Loop Check", "Overall"]

            for row_num, header in enumerate(pattern_headers, start=top_row):
                self.active_sheet.cell(row=row_num, column=18, value=header)

            total_headers = ["Total Positives", "True Positives", "False Positives",
                             "False Negatives", "Precision", "Recall"]

            for col_num, header in enumerate(total_headers, start=19):
                self.active_sheet.cell(row=top_row, column=col_num, value=header)
        except Exception as e:
            print(f"Error adding results table: {e}")

    def createHeaders(self, filepath: str):
        try:
            # load the workbook and choose active sheet. Let this be the sheet for OP-0
            self.workbook = load_workbook(filepath)
            self.active_sheet = self.workbook.active
            self.active_sheet.title = 'OP-0'

            # Iterate through all sheets in the workbook
            for sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]

                # Clear all rows and columns in the sheet
                sheet.delete_rows(1, sheet.max_row)
                sheet.delete_cols(1, sheet.max_column)

            # Merge the cells B1:D1 and set the text to "OP-0"
            self.active_sheet.merge_cells('B1:D1')
            self.active_sheet['B1'] = 'OP-0'

            # Merge the cells K1:P1 and set the text to "Lines Missed"
            self.active_sheet.merge_cells('K1:P1')
            self.active_sheet['K1'] = 'Lines Missed'

            # Add the OP-0 headers from cell A2 onwards
            op_headers = ["File Name", "Line No.", "Line", "Branch", "Constant Coding",
                          "Loop Check", "True Positives", "False Positive", "Comments"]

            for col_num, header in enumerate(op_headers, start=1):
                self.active_sheet.cell(row=2, column=col_num, value=header)

            lines_missed_headers = ["Line No.", "Line", "Branch", "Constant Coding", "Loop Check", "Comments"]

            for col_num, header in enumerate(lines_missed_headers, start=11):
                self.active_sheet.cell(row=2, column=col_num, value=header)

            # FOR TESTING PURPOSES
            # self.generate_resultsTable(2)

            # Create sheet for OP-1 and copy contents
            if 'OP-1' not in self.workbook.sheetnames:
                new_sheet = self.workbook.create_sheet('OP-1')
            else:
                new_sheet = self.workbook['OP-1']

            # Copy data, formatting, and other properties from the active sheet to the new sheet
            for row in self.active_sheet.iter_rows(min_row=1, max_row=self.active_sheet.max_row, min_col=1,
                                                   max_col=self.active_sheet.max_column):
                for cell in row:
                    new_cell = new_sheet[cell.coordinate]
                    new_cell.value = cell.value

            # Switch to OP-1 sheet
            self.active_sheet = self.workbook['OP-1']
            # Merge the cells B1:D1 and set the text to "OP-1"
            self.active_sheet.merge_cells('B1:D1')
            self.active_sheet['B1'] = 'OP-1'
            # Merge the cells K1:P1
            self.active_sheet.merge_cells('K1:P1')

            self.workbook.save(filepath)
            print(f"Headers and text successfully added to '{filepath}'.")

        except Exception as e:
            print(f"Error adding headers: {e}")

    def strip_file_name(self, file_name):
        # Define the patterns to remove from the front
        front_patterns = ['op_0_ai_', 'op_0_manual_', 'op_1_ai_', 'op_1_manual_']

        # Create a regex pattern for the front
        front_pattern = '|'.join(map(re.escape, front_patterns))
        front_regex = re.compile(f'^({front_pattern})')

        # Create a regex pattern for the back
        back_pattern = r'\.s\.out$'
        back_regex = re.compile(back_pattern)

        stripped_file_name = back_regex.sub('', front_regex.sub('', file_name))
        return stripped_file_name

    def fill_in_report(self, file_name):
        try:
            # Open the file and search for the target text
            with open(self.folder_path + file_name, 'r') as file:
                # what type of vulnerability
                fault_mode = None
                # whether lines are vulnerable code lines
                report_mode = False
                # collect all vulnerable lines
                current_line_buffer = []

                # set active sheet according to prefix of file
                if file_name[:4] == "op_0":
                    self.active_sheet = self.workbook['OP-0']
                elif file_name[:4] == "op_1":
                    self.active_sheet = self.workbook['OP-1']
                else:
                    raise ValueError(f"Cannot discern if file is of type OP-0 or OP-1")

                # locate point to write and write file name to report
                row_start = 3 if self.active_sheet.max_row == 2 else self.active_sheet.max_row + 2
                stripped_file_name = self.strip_file_name(file_name)
                self.active_sheet.cell(row=row_start, column=1, value=stripped_file_name)

                for line_no, line in enumerate(file, 1):
                    # set fault mode
                    if "BRANCH" in line:
                        fault_mode = "Branch"
                    elif "CONSTANT CODING" in line:
                        fault_mode = "ConstantCoding"
                    elif "LOOP_CHECK" in line:
                        fault_mode = "LoopCheck"

                    if line.strip().startswith("[Line #] [Opcode]"):
                        # set report mode true
                        report_mode = True
                        continue
                    elif line.strip().startswith("All vulnerable lines printed") and line:
                        # set report mode false
                        report_mode = False

                    if report_mode:
                        # if line is "", empty line in file and vulnerable lines should be written to report
                        if line.strip() == "" and len(current_line_buffer) > 0:
                            # writing line no. range
                            line_nos = "-".join([current_line_buffer[0].split()[0], current_line_buffer[-1].split()[0]]
                                                ) if len(current_line_buffer) > 1 else current_line_buffer[0].split()[0]
                            self.active_sheet.cell(row=row_start, column=2, value=line_nos)

                            # writing lines to report
                            lines = "\n".join([" ".join(s.split()[1:]) for s in current_line_buffer])
                            self.active_sheet.cell(row=row_start, column=3, value=lines)

                            # Entering TRUE into whichever vulnerability it is
                            if fault_mode == "Branch":
                                self.active_sheet.cell(row=row_start, column=4, value=True)
                            elif fault_mode == "ConstantCoding":
                                self.active_sheet.cell(row=row_start, column=5, value=True)
                            elif fault_mode == "LoopCheck":
                                self.active_sheet.cell(row=row_start, column=6, value=True)

                            row_start += 1
                            current_line_buffer.clear()
                        # else, it must be a vulnerable line. append
                        elif line.strip() != "":
                            current_line_buffer.append(line)

            self.workbook.save(self.workbook_path)
            print(f"\nSuccessfully added file {file_name[:-4]}")
        except Exception as e:
            print(f"\nError adding file {file_name[:-4]}: {e}")

    def locate_file_labelling(self, file_name):
        # start index
        start_row = 4
        while start_row <= self.active_labelling_sheet.max_row:
            if (self.active_labelling_sheet.cell(row=start_row, column=2).value is not None and
                    self.active_labelling_sheet.cell(row=start_row, column=2).value in file_name):
                return start_row
            start_row += 1

    def locate_file_analysis(self, file_name):
        # start index
        start_row = 3
        while start_row <= self.active_sheet.max_row:
            if (self.active_sheet.cell(row=start_row, column=1).value is not None and
                    self.active_sheet.cell(row=start_row, column=1).value in file_name):
                return start_row
            start_row += 1

    def get_patternType(self, sheet, row, col_start):
        for column_no in range(col_start, col_start + 3):
            if sheet.cell(row=row, column=column_no).value:
                return sheet.cell(row=2 if sheet.title in ['OP-0', 'OP-1'] else 3, column=column_no).value

    def check_precision_recall(self, file_name):
        try:
            # set active sheet according to prefix of file
            if file_name[:4] == "op_0":
                self.active_sheet = self.workbook['OP-0']
            elif file_name[:4] == "op_1":
                self.active_sheet = self.workbook['OP-1']
            else:
                raise ValueError(f"Cannot discern if file is of type OP-0 or OP-1")

            # locate the correct file labeling
            labelling_row_start = self.locate_file_labelling(file_name)

            # for op-0, checking if each vulnerability marked aligns with each vulnerability in labelling
            # locate start row
            analysis_row_start = self.locate_file_analysis(file_name)

            # loop through analysis vulnerabilities
            while True:
                analysis_line_nos = self.active_sheet.cell(row=analysis_row_start, column=2).value
                is_false_positive = True
                #labelling row no is first label for file
                labelling_row_no = labelling_row_start
                # if not blank row
                if analysis_line_nos:
                    while True:
                        labelling_line_nos = str(self.active_labelling_sheet.cell(row=labelling_row_no,
                                                                              column=8 if self.active_sheet.title == 'OP-0' else 13).value)
                        # check if the first digits match
                        if labelling_line_nos is not None and analysis_line_nos.split('-')[0] == labelling_line_nos.split('-')[0]:
                            # get pattern
                            analysis_patterntype = self.get_patternType(self.active_sheet, analysis_row_start, 4)
                            labelling_patterntype = self.get_patternType(self.active_labelling_sheet,labelling_row_no,
                                                                        10 if self.active_sheet.title == 'OP-0' else 15)
                            # if pattern types are equal, it must be the same vulnerability
                            if analysis_patterntype == labelling_patterntype:
                                is_false_positive = False
                                # mark True Positive as 'TRUE'
                                self.active_sheet.cell(row=analysis_row_start, column=7, value=True)
                                break

                        labelling_row_no += 1
                        # stop loop if all rows checked or next file detected
                        if labelling_row_no > self.active_labelling_sheet.max_row or self.active_labelling_sheet.cell(
                                row=labelling_row_no, column=2).value is not None:
                            break

                # if false positive, we mark False Positive as 'TRUE'
                if is_false_positive and analysis_line_nos is not None:
                    self.active_sheet.cell(row=analysis_row_start, column=8, value=True)

                analysis_row_start += 1
                # stop loop if all rows checked or next file detected
                if (analysis_row_start > self.active_sheet.max_row or
                        self.active_sheet.cell(row=analysis_row_start, column=1).value not in [None, ""]):
                    break

            # locate the correct file labeling
            labelling_row_start = self.locate_file_labelling(file_name)
            # for op-0, checking if each vulnerability marked aligns with each vulnerability in labelling
            # locate start row
            analysis_row_start = self.locate_file_analysis(file_name)
            # loop through labelling
            while True:
                # info to populate in case of false negative
                labelling_line_nos = str(self.active_labelling_sheet.cell(row=labelling_row_start,
                                                                          column=8 if self.active_sheet.title == 'OP-0' else 13).value)
                labelling_lines = self.active_labelling_sheet.cell(row=labelling_row_start,
                                                                   column=9 if self.active_sheet.title == 'OP-0' else 14).value
                labelling_patterntype = self.get_patternType(self.active_labelling_sheet, labelling_row_start,
                                                            10 if self.active_sheet.title == 'OP-0' else 15)
                is_false_negative = True
                # get first fault detected for file
                analysis_row_no = self.locate_file_analysis(file_name)
                # if not blank row
                if labelling_line_nos or labelling_line_nos != "None":
                    while True:
                        analysis_line_nos = str(self.active_sheet.cell(row=analysis_row_no, column=2).value)
                        # check if the first digits match
                        if analysis_line_nos is not None and labelling_line_nos.split('-')[0] == analysis_line_nos.split('-')[0]:
                            # get pattern
                            analysis_patterntype = self.get_patternType(self.active_sheet, analysis_row_no, 4)
                            # if pattern types are equal, it must be the same vulnerability
                            if analysis_patterntype == labelling_patterntype:
                                is_false_negative = False
                                break

                        analysis_row_no += 1
                        # stop loop if all rows checked or next file detected
                        if (analysis_row_no > self.active_sheet.max_row or
                                self.active_sheet.cell(row=analysis_row_no, column=1).value not in [None, ""]):
                            break

                if is_false_negative and labelling_line_nos and labelling_line_nos != "None":
                    self.active_sheet.cell(row=analysis_row_start, column=11, value=labelling_line_nos)
                    self.active_sheet.cell(row=analysis_row_start, column=12, value=labelling_lines)
                    self.active_sheet.cell(row=analysis_row_start, column=13 if labelling_patterntype == "Branch" else
                                           14 if labelling_patterntype == "Constant Coding" else 15
                                           if labelling_patterntype == "Loop Check" else 16, value=True)
                    analysis_row_start += 1

                labelling_row_start += 1
                # stop loop if all rows checked or next file detected
                if labelling_row_start > self.active_labelling_sheet.max_row or self.active_labelling_sheet.cell(
                        row=labelling_row_start, column=2).value is not None:
                    break

            self.workbook.save(self.workbook_path)
            print(f"Successfully generated precision and recall for {file_name}")
        except Exception as e:
            print(f"Error while finding precision and recall for {file_name}: {e}")

    def start_report(self):
        try:
            # check if folder exists
            if not os.path.exists(self.folder_path) or not os.path.isdir(self.folder_path):
                raise FileNotFoundError(f"Folder '{self.folder_path}' does not exist.")

            # get list of all files in the folder
            file_list = [f for f in os.listdir(self.folder_path) if os.path.isfile(os.path.join(self.folder_path, f))]

            # create workbook
            workbook_exists = self.createWorkbook(self.workbook_path)
            # now, populate headers
            if workbook_exists:
                self.createHeaders(self.workbook_path)

            # Fill in all vulnerable lines detected
            for file in file_list:
                self.fill_in_report(file)
                self.check_precision_recall(file)

        except FileNotFoundError as e:
            print(f"Error: {e}")
        except Exception as e:
            print(f"An unexpected error occurred: {e}")

# Define the command line arguments
parser = argparse.ArgumentParser(description='Get Analysis Report')
parser.add_argument('-f', '--input_path', required=True, help='Path to the input directory')
parser.add_argument('-w', '--output_path', required=True, help='Path to the output directory')

# Parse the command line arguments
args = parser.parse_args()

# Use the arguments to initialize your Analysis_Report
Analysis_Report = GetAnalysis(args.input_path, args.output_path)
Analysis_Report.start_report()

